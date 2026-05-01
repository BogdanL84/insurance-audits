"""Read the Template_Contract_Analysis.xlsx and dump every cell from every tab to stdout."""
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(r"C:\Users\Bogdan\Documents\insurance-audits\knowledge-base\methodology\Template_Contract_Analysis.xlsx")
wb = load_workbook(str(XLSX), data_only=True)

print(f"=== Workbook: {XLSX.name} ===")
print(f"Sheet names: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}\nSHEET: {sheet_name!r}  ({ws.max_row} rows × {ws.max_column} cols)\n{'='*80}")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Skip rows that are entirely empty
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        cells = []
        for c in row:
            if c is None:
                cells.append("")
            elif isinstance(c, str):
                cells.append(c.replace("\n", " ¶ ").strip())
            else:
                cells.append(str(c))
        # Trim trailing empty cells
        while cells and cells[-1] == "":
            cells.pop()
        print(f"  R{row_idx:3d}: {' | '.join(cells)}")
